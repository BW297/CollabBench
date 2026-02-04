#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import math
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter


def _safe_float(x: Any) -> Optional[float]:
    try:
        if x is None or (isinstance(x, float) and (math.isnan(x) or math.isinf(x))):
            return None
        if isinstance(x, str) and not x.strip():
            return None
        v = float(x)
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    except Exception:
        return None


def _task_sort_key_coela(task: str) -> Tuple[int, str]:
    m = re.match(r"^(\d+)_", str(task))
    if m:
        return int(m.group(1)), str(task)
    return 10**9, str(task)


PROAGENT_TASK_ORDER = [
    "asymmetric_advantages",
    "coordination_ring",
    "counter_circuit",
    "cramped_room",
    "forced_coordination",
]


def _task_sort_key_proagent(task: str) -> Tuple[int, str]:
    t = str(task)
    t2 = t
    if t2.startswith("-1_"):
        t2 = t2[3:]
    try:
        return PROAGENT_TASK_ORDER.index(t2), t
    except ValueError:
        return 10**9, t


def _copy_cell(dst: Cell, src: Cell) -> None:
    dst.value = src.value
    dst.number_format = src.number_format
    dst.font = Font(
        name=src.font.name,
        size=src.font.size,
        bold=src.font.bold,
        italic=src.font.italic,
        vertAlign=src.font.vertAlign,
        underline=src.font.underline,
        strike=src.font.strike,
        color=src.font.color,
    )
    dst.fill = PatternFill(
        fill_type=src.fill.fill_type,
        start_color=src.fill.start_color,
        end_color=src.fill.end_color,
        patternType=src.fill.patternType,
    )
    dst.border = Border(
        left=src.border.left,
        right=src.border.right,
        top=src.border.top,
        bottom=src.border.bottom,
        diagonal=src.border.diagonal,
        diagonal_direction=src.border.diagonal_direction,
        outline=src.border.outline,
        vertical=src.border.vertical,
        horizontal=src.border.horizontal,
    )
    dst.alignment = Alignment(
        horizontal=src.alignment.horizontal,
        vertical=src.alignment.vertical,
        text_rotation=src.alignment.text_rotation,
        wrap_text=src.alignment.wrap_text,
        shrink_to_fit=src.alignment.shrink_to_fit,
        indent=src.alignment.indent,
    )
    dst.protection = Protection(locked=src.protection.locked, hidden=src.protection.hidden)


def _copy_block(
    *,
    src_ws,
    dst_ws,
    src_top: int,
    src_left: int,
    src_bottom: int,
    src_right: int,
    dst_top: int,
    dst_left: int,
) -> None:
    n_rows = src_bottom - src_top + 1
    n_cols = src_right - src_left + 1
    for r in range(n_rows):
        sr = src_top + r
        dr = dst_top + r
        if src_ws.row_dimensions[sr].height is not None:
            dst_ws.row_dimensions[dr].height = src_ws.row_dimensions[sr].height
        for c in range(n_cols):
            sc = src_left + c
            dc = dst_left + c
            _copy_cell(dst_ws.cell(dr, dc), src_ws.cell(sr, sc))


def _copy_column_widths(src_ws, dst_ws, *, left: int, right: int) -> None:
    for c in range(left, right + 1):
        col = get_column_letter(c)
        w = src_ws.column_dimensions[col].width
        if w is not None:
            dst_ws.column_dimensions[col].width = w


def _copy_merges_with_offset(src_ws, dst_ws, *, src_row_min: int, src_row_max: int, row_offset: int) -> None:
    # Copy merged ranges that are fully within [src_row_min, src_row_max] rows.
    for rng in list(src_ws.merged_cells.ranges):
        if rng.min_row < src_row_min or rng.max_row > src_row_max:
            continue
        dst_ws.merge_cells(
            start_row=rng.min_row + row_offset,
            start_column=rng.min_col,
            end_row=rng.max_row + row_offset,
            end_column=rng.max_col,
        )


def _copy_merges_with_row_col_offset(
    src_ws,
    dst_ws,
    *,
    src_row_min: int,
    src_row_max: int,
    src_col_min: int,
    src_col_max: int,
    row_offset: int,
    col_offset: int,
) -> None:
    for rng in list(src_ws.merged_cells.ranges):
        if rng.min_row < src_row_min or rng.max_row > src_row_max:
            continue
        if rng.min_col < src_col_min or rng.max_col > src_col_max:
            continue
        dst_ws.merge_cells(
            start_row=rng.min_row + row_offset,
            start_column=rng.min_col + col_offset,
            end_row=rng.max_row + row_offset,
            end_column=rng.max_col + col_offset,
        )


@dataclass(frozen=True)
class RowSpec:
    method_label: str
    llm_display: str
    # how to find the row in CSV
    csv_method: str
    csv_llm: str
    csv_model_key: Optional[str] = None


def _get_row(df: pd.DataFrame, spec: RowSpec, task: str) -> Optional[pd.Series]:
    sub = df[df["Task"] == task]
    sub = sub[sub["Method"] == spec.csv_method]
    sub = sub[sub["LLMs"] == spec.csv_llm]
    if spec.csv_model_key is not None and "ModelKey" in sub.columns:
        sub = sub[sub["ModelKey"] == spec.csv_model_key]
    if len(sub) == 0:
        return None
    # Expect unique.
    return sub.iloc[0]


def _fill_data_row(
    ws,
    row_idx: int,
    *,
    method: Optional[str],
    llm: str,
    data: Optional[pd.Series],
    col_offset: int = 0,
) -> None:
    # col_offset=0 matches the template (Method=A, LLMs=B, metrics start at C).
    # If we insert Task as the first column, we can set col_offset=1 (Method=B, LLMs=C, metrics start at D).
    c1 = ws.cell(row_idx, 1 + col_offset)
    if not isinstance(c1, MergedCell):
        c1.value = method
    ws.cell(row_idx, 2 + col_offset).value = llm
    if data is None:
        # Clear numeric cells C-N
        for c in range(3 + col_offset, 15 + col_offset):
            ws.cell(row_idx, c).value = None
        return
    cols = [
        ("Agent 1 Score", 3 + col_offset),
        ("Agent 2 Score", 4 + col_offset),
        ("Agent 1 Std.", 5 + col_offset),
        ("Agent 2 Std.", 6 + col_offset),
        ("Agent 1 #Tokens(k)", 7 + col_offset),
        ("Agent 2 #Tokens(k)", 8 + col_offset),
        ("Agent 1 Helpfulness", 9 + col_offset),
        ("Agent 2 Helpfulness", 10 + col_offset),
        ("Agent 1 Trustfulness", 11 + col_offset),
        ("Agent 2 Trustfulness", 12 + col_offset),
        ("Agent 1 Empathy", 13 + col_offset),
        ("Agent 2 Empathy", 14 + col_offset),
    ]
    for col_name, col_idx in cols:
        v = data.get(col_name) if col_name in data else None
        ws.cell(row_idx, col_idx).value = _safe_float(v)


def _set_rel_improv_formulas(ws, rel_row: int, base_row: int, ours_row: int) -> None:
    # Match the ICML template: percent format already copied.
    # Efficiency (lower better): (base - ours) / base
    ws.cell(rel_row, 3).value = f"=(C{base_row}-C{ours_row})/C{base_row}"
    ws.cell(rel_row, 4).value = f"=(D{base_row}-D{ours_row})/D{base_row}"
    ws.cell(rel_row, 5).value = f"=-(E{ours_row}-E{base_row})/E{base_row}"
    ws.cell(rel_row, 6).value = f"=-(F{ours_row}-F{base_row})/F{base_row}"
    ws.cell(rel_row, 7).value = f"=-(G{ours_row}-G{base_row})/G{base_row}"
    ws.cell(rel_row, 8).value = f"=-(H{ours_row}-H{base_row})/H{base_row}"
    # Affective (higher better): (ours - base) / base
    ws.cell(rel_row, 9).value = f"=(I{ours_row}-I{base_row})/I{base_row}"
    ws.cell(rel_row, 10).value = f"=(J{ours_row}-J{base_row})/J{base_row}"
    ws.cell(rel_row, 11).value = f"=(K{ours_row}-K{base_row})/K{base_row}"
    ws.cell(rel_row, 12).value = f"=(L{ours_row}-L{base_row})/L{base_row}"
    ws.cell(rel_row, 13).value = f"=(M{ours_row}-M{base_row})/M{base_row}"
    ws.cell(rel_row, 14).value = f"=(N{ours_row}-N{base_row})/N{base_row}"


def _set_rel_improv_formulas_shifted(
    ws,
    *,
    rel_row: int,
    base_row: int,
    ours_row: int,
    col_offset: int,
) -> None:
    # With Task as first column, metrics shift right by 1:
    # Score columns C/D become D/E, ... N becomes O.
    def col(n: int) -> str:
        return get_column_letter(n + col_offset)

    # Efficiency (lower better): (base - ours) / base
    ws.cell(rel_row, 3 + col_offset).value = f"=({col(3)}{base_row}-{col(3)}{ours_row})/{col(3)}{base_row}"
    ws.cell(rel_row, 4 + col_offset).value = f"=({col(4)}{base_row}-{col(4)}{ours_row})/{col(4)}{base_row}"
    ws.cell(rel_row, 5 + col_offset).value = f"=-({col(5)}{ours_row}-{col(5)}{base_row})/{col(5)}{base_row}"
    ws.cell(rel_row, 6 + col_offset).value = f"=-({col(6)}{ours_row}-{col(6)}{base_row})/{col(6)}{base_row}"
    ws.cell(rel_row, 7 + col_offset).value = f"=-({col(7)}{ours_row}-{col(7)}{base_row})/{col(7)}{base_row}"
    ws.cell(rel_row, 8 + col_offset).value = f"=-({col(8)}{ours_row}-{col(8)}{base_row})/{col(8)}{base_row}"
    # Affective (higher better): (ours - base) / base
    ws.cell(rel_row, 9 + col_offset).value = f"=({col(9)}{ours_row}-{col(9)}{base_row})/{col(9)}{base_row}"
    ws.cell(rel_row, 10 + col_offset).value = f"=({col(10)}{ours_row}-{col(10)}{base_row})/{col(10)}{base_row}"
    ws.cell(rel_row, 11 + col_offset).value = f"=({col(11)}{ours_row}-{col(11)}{base_row})/{col(11)}{base_row}"
    ws.cell(rel_row, 12 + col_offset).value = f"=({col(12)}{ours_row}-{col(12)}{base_row})/{col(12)}{base_row}"
    ws.cell(rel_row, 13 + col_offset).value = f"=({col(13)}{ours_row}-{col(13)}{base_row})/{col(13)}{base_row}"
    ws.cell(rel_row, 14 + col_offset).value = f"=({col(14)}{ours_row}-{col(14)}{base_row})/{col(14)}{base_row}"


def _build_sheet(
    *,
    dst_ws,
    src_ws,
    title_prefix: str,
    df: pd.DataFrame,
    tasks: List[str],
    row_specs: List[RowSpec],
    baseline_row_in_block: int,
    ours_row_in_block: int,
    block_height: int = 10,
    spacer: int = 1,
) -> None:
    # Copy column widths from template A-N.
    _copy_column_widths(src_ws, dst_ws, left=1, right=14)
    # Copy merges from template block rows 1-10 once per block with offset.
    for i, task in enumerate(tasks):
        top = 1 + i * (block_height + spacer)
        _copy_block(
            src_ws=src_ws,
            dst_ws=dst_ws,
            src_top=1,
            src_left=1,
            src_bottom=block_height,
            src_right=14,
            dst_top=top,
            dst_left=1,
        )
        _copy_merges_with_offset(src_ws, dst_ws, src_row_min=1, src_row_max=block_height, row_offset=top - 1)

        dst_ws.cell(top, 1).value = f"{title_prefix}: {task}"

        # Fill data rows 5-9 of the block.
        for j, spec in enumerate(row_specs):
            row_in_block = 5 + j
            dest_row = top + (row_in_block - 1)
            data = _get_row(df, spec, task)
            method_cell: Optional[str] = spec.method_label
            # Match template grouping: leave Method blank for subsequent baseline LLM rows under the same method label.
            if j > 0 and spec.method_label == row_specs[0].method_label:
                method_cell = None
            _fill_data_row(dst_ws, dest_row, method_cell, spec.llm_display, data)

        # Set rel improv formulas.
        rel_row = top + (10 - 1)
        base_row = top + (baseline_row_in_block - 1)
        ours_row = top + (ours_row_in_block - 1)
        _set_rel_improv_formulas(dst_ws, rel_row, base_row, ours_row)


def _copy_header_from_template(src_ws, dst_ws, *, add_task_col: bool) -> None:
    if add_task_col:
        # Big-table mode: Task is the FIRST column, so shift template block right by 1.
        # Copy header A-N -> B-O
        _copy_column_widths(src_ws, dst_ws, left=1, right=14)
        for c in range(1, 15):
            col_src = get_column_letter(c)
            col_dst = get_column_letter(c + 1)
            dst_ws.column_dimensions[col_dst].width = dst_ws.column_dimensions[col_src].width
        for r in range(1, 5):
            if src_ws.row_dimensions[r].height is not None:
                dst_ws.row_dimensions[r].height = src_ws.row_dimensions[r].height
            for c in range(1, 15):
                _copy_cell(dst_ws.cell(r, c + 1), src_ws.cell(r, c))
        _copy_merges_with_row_col_offset(src_ws, dst_ws, src_row_min=1, src_row_max=4, src_col_min=1, src_col_max=14, row_offset=0, col_offset=1)

        # Task column A: style from template column A.
        dst_ws.column_dimensions[get_column_letter(1)].width = 22
        for r in range(1, 5):
            _copy_cell(dst_ws.cell(r, 1), src_ws.cell(r, 1))
        dst_ws.cell(4, 1).value = "Task"

        # Expand "Metric" header to include Task as the first column.
        # Template has A2:B3 merged; shifted copy yields B2:C3 merged.
        try:
            dst_ws.unmerge_cells("B2:C3")
        except Exception:
            pass
        dst_ws.merge_cells("A2:C3")
        dst_ws.cell(2, 1).value = "Metric"
        # Clear the old top-left text if it exists (only A2 matters after merge).
        for cell in ("B2", "C2", "A3", "B3", "C3"):
            try:
                dst_ws[cell].value = None
            except Exception:
                pass
        return

    # Non-shifted mode: copy rows 1-4 (A-N) exactly.
    _copy_column_widths(src_ws, dst_ws, left=1, right=14)
    _copy_block(src_ws=src_ws, dst_ws=dst_ws, src_top=1, src_left=1, src_bottom=4, src_right=14, dst_top=1, dst_left=1)
    _copy_merges_with_offset(src_ws, dst_ws, src_row_min=1, src_row_max=4, row_offset=0)


def _write_big_table(
    *,
    dst_ws,
    src_ws,
    title: str,
    df: pd.DataFrame,
    tasks: List[str],
    row_specs: List[RowSpec],
    baseline_method_label: str,
) -> None:
    # Header once.
    _copy_header_from_template(src_ws, dst_ws, add_task_col=True)
    # Title lives in merged B1:O1 (shifted template header).
    dst_ws.cell(1, 2).value = title
    dst_ws.cell(1, 1).value = None

    start_row = 5
    for task in tasks:
        task_top = start_row
        for j, spec in enumerate(row_specs):
            r = start_row + j
            src_r = 5 + j  # template data row
            # Copy styling from template A-N -> B-O.
            for c in range(1, 15):
                _copy_cell(dst_ws.cell(r, c + 1), src_ws.cell(src_r, c))
            if src_ws.row_dimensions[src_r].height is not None:
                dst_ws.row_dimensions[r].height = src_ws.row_dimensions[src_r].height
            # Task column A style.
            _copy_cell(dst_ws.cell(r, 1), src_ws.cell(src_r, 1))

            data = _get_row(df, spec, task)
            _fill_data_row(dst_ws, r, method=spec.method_label, llm=spec.llm_display, data=data, col_offset=1)

        # Add Rel. Improv. row (template row 10) after the 5 data rows.
        rel_row = start_row + len(row_specs)
        for c in range(1, 15):
            _copy_cell(dst_ws.cell(rel_row, c + 1), src_ws.cell(10, c))
        if src_ws.row_dimensions[10].height is not None:
            dst_ws.row_dimensions[rel_row].height = src_ws.row_dimensions[10].height
        _copy_cell(dst_ws.cell(rel_row, 1), src_ws.cell(10, 1))
        # merge label cells like template A10:B10 -> B10:C10 (shifted)
        dst_ws.merge_cells(start_row=rel_row, start_column=2, end_row=rel_row, end_column=3)
        dst_ws.cell(rel_row, 2).value = "Rel. Improv."
        _set_rel_improv_formulas_shifted(
            dst_ws,
            rel_row=rel_row,
            base_row=start_row + 3,  # baseline row: 4th data row in group
            ours_row=start_row + 4,  # ours row: 5th data row in group
            col_offset=1,
        )

        # Merge Task label column A across 6 rows (5 data + rel).
        dst_ws.merge_cells(start_row=task_top, start_column=1, end_row=rel_row, end_column=1)
        dst_ws.cell(task_top, 1).value = task

        # Merge baseline method label across the first 4 rows (like template).
        dst_ws.merge_cells(start_row=task_top, start_column=2, end_row=task_top + 3, end_column=2)
        dst_ws.cell(task_top, 2).value = baseline_method_label
        # SynerMate row label in 5th row.
        dst_ws.cell(task_top + 4, 2).value = row_specs[4].method_label

        start_row += len(row_specs) + 1


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Build ICML-style per-task summary xlsx from *_tasks_raw.csv.")
    ap.add_argument("--template", required=True, help="Path to the existing ICML(.xlsx) template.")
    ap.add_argument("--coela-tasks-csv", required=True)
    ap.add_argument("--proagent-tasks-csv", required=True)
    ap.add_argument("--out", required=True, help="Output xlsx path.")
    ap.add_argument("--coela-layout", choices=["blocks", "big"], default="big")
    ap.add_argument("--proagent-layout", choices=["blocks", "big"], default="big")
    args = ap.parse_args(argv)

    template = Path(os.path.expanduser(args.template)).resolve()
    wb_t = load_workbook(template)
    src_ws = wb_t.active

    df_coela = pd.read_csv(args.coela_tasks_csv)
    df_pro = pd.read_csv(args.proagent_tasks_csv)

    coela_tasks = sorted(df_coela["Task"].dropna().unique().tolist(), key=_task_sort_key_coela)
    pro_tasks = sorted(df_pro["Task"].dropna().unique().tolist(), key=_task_sort_key_proagent)

    coela_specs = [
        RowSpec("CoELA", "GPT-5.2", "CoELA", "GPT"),
        RowSpec("CoELA", "DeepSeek-V3.1", "CoELA", "DeepSeek-V3.1"),
        RowSpec("CoELA", "Qwen2.5-72B-Instruct", "CoELA", "Qwen2.5-72B-Instruct"),
        RowSpec("CoELA", "Qwen2.5-7B-Instruct", "CoELA", "Qwen2.5-7B-Instruct", "qwen7b"),
        RowSpec("SynerMate", "Qwen2.5-7B-Instruct", "SynerMate", "Qwen2.5-7B-Instruct", "qwen7brl"),
    ]

    pro_specs = [
        RowSpec("ProAgent", "GPT-5.2", "ProAgent", "GPT"),
        RowSpec("ProAgent", "DeepSeek-V3.1", "ProAgent", "DeepSeek-V3.1"),
        RowSpec("ProAgent", "Qwen2.5-72B-Instruct", "ProAgent", "Qwen2.5-72B-Instruct"),
        RowSpec("ProAgent", "Qwen2.5-7B-Instruct", "ProAgent", "Qwen2.5-7B-Instruct", "qwen7b"),
        RowSpec("SynerMate", "ours", "ProAgent", "ours", "ours"),
    ]

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("CoELA-PerTask")
    if args.coela_layout == "blocks":
        _build_sheet(
            dst_ws=ws1,
            src_ws=src_ws,
            title_prefix="CWAH-MultiPlayer",
            df=df_coela,
            tasks=coela_tasks,
            row_specs=coela_specs,
            baseline_row_in_block=8,
            ours_row_in_block=9,
        )
    else:
        _write_big_table(
            dst_ws=ws1,
            src_ws=src_ws,
            title="CWAH-MultiPlayer",
            df=df_coela,
            tasks=coela_tasks,
            row_specs=coela_specs,
            baseline_method_label="CoELA",
        )

    ws2 = wb.create_sheet("ProAgent-PerTask")
    if args.proagent_layout == "blocks":
        _build_sheet(
            dst_ws=ws2,
            src_ws=src_ws,
            title_prefix="Cook-MultiPlayer",
            df=df_pro,
            tasks=pro_tasks,
            row_specs=pro_specs,
            baseline_row_in_block=8,
            ours_row_in_block=9,
        )
    else:
        _write_big_table(
            dst_ws=ws2,
            src_ws=src_ws,
            title="Cook-MultiPlayer",
            df=df_pro,
            tasks=pro_tasks,
            row_specs=pro_specs,
            baseline_method_label="ProAgent",
        )

    out = Path(os.path.expanduser(args.out)).resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
